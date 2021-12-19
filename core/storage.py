# -*- coding:utf-8 -*-
__author__ = 'linxinloningg'

import pandas as pd
from csv import reader
from openpyxl import load_workbook, Workbook
import os
from datetime import datetime


class Dataoperation:
    """

    """

    def __init__(self):
        pass

    @staticmethod
    def write2csv(datadict, filepath):
        """

        :param datadict:
        :param filepath:
        :return:
        """
        pd.DataFrame(datadict, index=[filepath]).to_csv(filepath)
        return True

    @staticmethod
    def read2csv(filepath):
        """

        :param filepath:
        :return:
        """
        datadict = {}

        csv = reader(open("{}.csv".format(filepath), 'r', encoding='UTF-8'))

        codes = next(csv)
        codes.pop(0)
        values = next(csv)
        values.pop(0)

        for i in range(len(codes)):
            datadict[codes[i]] = values[i]

        return datadict

    @staticmethod
    def write2txt(content, filename):
        """
        # 如写入出错，可选字符串处理
        trantab = str.maketrans("?*/\|.:><", "         ")
        content = content.translate(trantab)
        :param content:
        :param filename:
        :return:
        """
        if filename is None:
            return False
        else:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(content)
            return True

    @staticmethod
    def read2txt(filename):
        """

        :param filename:
        :return:
        """
        if filename is None:
            return False
        else:
            with open(filename, 'r', encoding='utf-8') as f:
                content = f.read()
            return content


class Exceloperation:
    """

    """

    def __init__(self, filename):
        """

        :param filename:
        """
        self.filename = filename

    def write(self, data):
        """

        :param data:
        :return:
        """
        if os.path.exists(self.filename):
            workbook = load_workbook(self.filename)
        else:
            workbook = Workbook()

        # 获取数据属性，查看是否存在此表单，若不存在，则创建
        formname = list(data.keys())[-1]
        if formname not in workbook.sheetnames:
            newSheet = workbook.create_sheet(formname)
            newSheet.title = formname
        worksheet = workbook[formname]

        Header = list(data.keys())
        Header.reverse()
        rows = [
            [Header[0], data[Header[0]]],
        ]
        for i in range(1, len(Header)):
            rows.append([Header[i]])
            if len(list(data[Header[i]].keys())) != 0:
                for code in list(data[Header[i]].keys()):
                    temp = [code, data[Header[i]][code][0], data[Header[i]][code][1]]
                    rows.append(temp)
            else:
                rows.append(['None'])
            rows.append(['*' * 20])

        for row in rows:
            worksheet.append(row)

        workbook.save(self.filename)
        return True

    def read(self):
        """

        :return:
        """
        if os.path.exists(self.filename) is False:

            return
        else:
            workbook = load_workbook(self.filename)

            # 返还的买卖数据dict
            datadict = {}
            datadict.setdefault('buy', list())
            datadict.setdefault('sell', list())
            # 输出工作簿中所有表单名字

            # 构建表单列表
            sheetlist = list()
            for sheetname in workbook.sheetnames:
                sheetlist.append(workbook.get_sheet_by_name(sheetname))

            for sheet in sheetlist:
                # 遍历表单找到buy数据所在
                buypointlist = list()
                for i in range(1, sheet.max_row + 1):
                    for j in range(1, sheet.max_column + 1):
                        if sheet.cell(row=i, column=j).value == 'buy':
                            buypointlist.append([i, j])
                # 遍历表单找到sell数据所在
                sellpointlist = list()
                for i in range(1, sheet.max_row + 1):
                    for j in range(1, sheet.max_column + 1):
                        if sheet.cell(row=i, column=j).value == 'sell':
                            sellpointlist.append([i, j])
                # 遍历表单找到*数据所在
                pointlist = list()
                for i in range(1, sheet.max_row + 1):
                    for j in range(1, sheet.max_column + 1):
                        if sheet.cell(row=i, column=j).value == '*' * 20:
                            pointlist.append([i, j])
                # 排除空表单
                if len(pointlist) > 0:
                    # 最新的一个买数据开始点
                    buy = buypointlist[-1][0] + 1
                    stop = pointlist[-1][0] - 1
                    cell_range = sheet["{}:{}".format(buy, stop)]
                    # 排除空数据
                    if buy != stop:
                        totalbuydata = list()
                        for row_range in cell_range:  # 先行再列
                            buydata = list()
                            for cell in row_range:
                                buydata.append(cell.value)
                            totalbuydata.append(buydata)
                    else:
                        totalbuydata = list()

                    datadict['buy'].extend(totalbuydata)

                    # 最新的一个卖数据开始点
                    sell = sellpointlist[-1][0] + 1
                    stop = pointlist[-2][0] - 1
                    cell_range = sheet["{}:{}".format(sell, stop)]
                    # 排除空数据
                    if sell != stop:
                        totalselldata = list()
                        for row_range in cell_range:  # 先行再列
                            selldata = list()
                            for cell in row_range:
                                selldata.append(cell.value)
                            totalselldata.append(selldata)
                    else:
                        totalselldata = list()

                    datadict['sell'].extend(totalselldata)
            return datadict

    def updateornot(self):
        """

        :return:
        """
        if os.path.exists(self.filename) is False:

            return
        else:
            workbook = load_workbook(self.filename)
            # 输出工作簿中所有表单名字

            # 构建表单列表
            sheetlist = list()
            for sheetname in workbook.sheetnames:
                sheetlist.append(workbook.get_sheet_by_name(sheetname))

            for sheet in sheetlist:
                # 遍历表单找到日期数据所在
                for i in range(1, sheet.max_row + 1):
                    for j in range(1, sheet.max_column + 1):
                        if sheet.cell(row=i, column=j).value == datetime.now().strftime("%Y/%m/%d"):

                            return True

                        else:

                            return False
